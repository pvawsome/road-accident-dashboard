from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
GUIDE_PATH = (
    PROJECT_ROOT
    / "data"
    / "raw"
    / "road_safety_data_guide_2024.xlsx"
)


def main() -> None:
    if not GUIDE_PATH.exists():
        raise FileNotFoundError(f"Data guide not found: {GUIDE_PATH}")

    workbook = load_workbook(
        GUIDE_PATH,
        read_only=True,
        data_only=True,
    )

    print("=" * 70)
    print("WORKBOOK SHEETS")
    print("=" * 70)

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        print(f"\nSheet: {sheet_name}")
        print(f"Rows: {worksheet.max_row}")
        print(f"Columns: {worksheet.max_column}")

        print("First eight rows:")

        for row in worksheet.iter_rows(
            min_row=1,
            max_row=min(8, worksheet.max_row),
            values_only=True,
        ):
            print(row)

        print("-" * 70)


if __name__ == "__main__":
    main()
